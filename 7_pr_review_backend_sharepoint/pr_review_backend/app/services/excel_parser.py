"""
Parsing Excel robuste (Lot 5, §6.2.5).

Les fichiers importés sont hétérogènes : la colonne de texte n'est pas toujours en
A, les booléens sont parfois stockés en texte ('True', 'x', 'oui'), et les
colonnes de statut peuvent se trouver n'importe où. Ce module :

  - détecte dynamiquement la colonne de TEXTE (celle dont les chaînes sont les
    plus longues en moyenne) ;
  - pour chaque ligne, retient la meilleure valeur de STATUT parmi les autres
    colonnes (première cellule interprétable comme OK/KO/Partiel/N-A).

Ce comportement reproduit la correction éprouvée sur les fichiers réels du
prototype (texte en colonne B, booléens 'True', statuts à partir de la colonne C).
"""
from dataclasses import dataclass


@dataclass
class ParsedRow:
    text: str
    status_raw: str | None


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_workbook(data: bytes, status_detector=None) -> list[ParsedRow]:
    """
    Parse un classeur Excel (bytes) en lignes {texte, valeur de statut brute}.

    `status_detector` (optionnel) : fonction str -> ItemStatus|None servant à
    choisir, parmi les colonnes non-texte, celle qui ressemble le plus à un statut.
    """
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows: list[list] = []
    for r in ws.iter_rows(values_only=True):
        if r is None:
            continue
        if all(_cell_str(c) == "" for c in r):
            continue
        rows.append(list(r))

    if not rows:
        return []

    n_cols = max(len(r) for r in rows)
    for r in rows:
        r.extend([None] * (n_cols - len(r)))

    # --- Détection de la colonne de texte : plus longues chaînes en moyenne. ---
    # On ignore la première ligne (souvent un en-tête) pour l'estimation.
    body = rows[1:] if len(rows) > 1 else rows
    avg_len = [0.0] * n_cols
    for c in range(n_cols):
        lengths = [len(_cell_str(r[c])) for r in body]
        avg_len[c] = sum(lengths) / len(lengths) if lengths else 0.0
    text_col = max(range(n_cols), key=lambda c: avg_len[c])

    # Détecte si la 1re ligne est un en-tête (peu de texte long, pas de statut).
    def _looks_like_header(row) -> bool:
        first = _cell_str(row[text_col]).lower()
        return first in ("", "règle", "regle", "rule", "libellé", "libelle", "description")

    start = 1 if rows and _looks_like_header(rows[0]) else 0

    parsed: list[ParsedRow] = []
    for r in rows[start:]:
        text = _cell_str(r[text_col])
        if not text:
            continue
        # Meilleure valeur de statut parmi les autres colonnes.
        status_raw = None
        for c in range(n_cols):
            if c == text_col:
                continue
            candidate = _cell_str(r[c])
            if not candidate:
                continue
            if status_detector is not None:
                if status_detector(candidate) is not None:
                    status_raw = candidate
                    break
            else:
                status_raw = candidate
                break
        parsed.append(ParsedRow(text=text, status_raw=status_raw))
    return parsed
