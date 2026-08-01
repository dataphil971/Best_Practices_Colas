"""
Routes de gestion des revues (Lot 3).

Couvre le cycle de vie complet d'une revue côté auteur :
  - création (fige le snapshot des versions courantes) ;
  - liste filtrable et recherche ;
  - vue détaillée groupée par catégorie (items unset inclus) ;
  - mise à jour statut / progress / remédiation d'un item, avec recalcul du score ;
  - renommage, changement de statut, suppression ;
  - export Excel formaté (colonnes OK/KO/Partiel/N-A cochées).

Règle d'accès (visibilité stricte, conforme à la spec) :
  - l'AUTEUR voit et modifie ses propres revues ;
  - un ADMIN voit toutes les revues (droits complets) ;
  - le partage ciblé vers des reviewers précis arrive au Lot 4.
"""
import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.auth.deps import get_current_user
from app.models.user import User
from app.models.enums import UserRole, ChecklistType, ReviewStatus, ItemStatus
from app.models.review import Review, ReviewItem
from app.schemas.review import (
    ReviewCreate,
    ReviewUpdate,
    ReviewItemUpdate,
    ReviewSummary,
    ReviewDetail,
    ReviewItemOut,
)
from app.services import review as svc
from app.services.review_serializers import review_detail, review_summary

router = APIRouter(prefix="/reviews", tags=["revues"])


# --------------------------------------------------------------------------
# Accès
# --------------------------------------------------------------------------
def _get_review_or_404(db: Session, review_id: uuid.UUID) -> Review:
    review = db.get(Review, review_id)
    if review is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Revue introuvable.")
    return review


def _assert_can_view(review: Review, user: User, db: Session) -> None:
    if user.role == UserRole.admin or review.author_id == user.id:
        return
    # Visibilité reviewer : uniquement si un lien de partage actif le cible.
    from app.services.validation import can_user_view_review
    if can_user_view_review(db, review=review, user=user):
        return
    raise HTTPException(status.HTTP_403_FORBIDDEN, "Accès à cette revue non autorisé.")


def _assert_is_author(review: Review, user: User) -> None:
    if review.author_id != user.id:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "Seul l'auteur peut modifier cette revue."
        )


# --------------------------------------------------------------------------
# Création
# --------------------------------------------------------------------------
@router.post("", response_model=ReviewDetail, status_code=status.HTTP_201_CREATED)
def create_review(
    payload: ReviewCreate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Créer une revue : fige un item par règle active du référentiel choisi."""
    review = svc.create_review(
        db,
        author=user,
        report_name=payload.report_name,
        checklist_type=payload.checklist_type,
    )
    db.commit()
    db.refresh(review)
    return review_detail(db, review)


# --------------------------------------------------------------------------
# Liste + recherche
# --------------------------------------------------------------------------
@router.get("", response_model=list[ReviewSummary])
def list_reviews(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    q: str | None = None,
    status_filter: str | None = None,
    type: str | None = None,
):
    """
    Liste des revues visibles.

    - `admin`    → toutes les revues ;
    - `user`     → les siennes (celles dont il est l'auteur) ;
    - `reviewer` → les siennes ET celles qui lui ont été partagées via un lien
      actif le ciblant. Aucun accès de large périmètre.
    - `q`    → recherche sur le nom de rapport.
    - `status_filter` (`?status=`) et `type` filtrent la liste.
    """
    query = select(Review)
    if user.role != UserRole.admin:
        from app.services.validation import list_visible_review_ids_for_reviewer
        shared_ids = list_visible_review_ids_for_reviewer(db, user)
        if shared_ids:
            query = query.where(
                (Review.author_id == user.id) | (Review.id.in_(shared_ids))
            )
        else:
            query = query.where(Review.author_id == user.id)
    if q:
        query = query.where(Review.report_name.ilike(f"%{q}%"))
    if status_filter:
        try:
            query = query.where(Review.status == ReviewStatus(status_filter))
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Statut inconnu.")
    if type:
        try:
            query = query.where(Review.checklist_type == ChecklistType(type))
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Type de référentiel inconnu.")

    query = query.order_by(Review.created_at.desc())
    reviews = db.scalars(query).all()
    return [review_summary(db, r) for r in reviews]


# --------------------------------------------------------------------------
# Détail
# --------------------------------------------------------------------------
@router.get("/{review_id}", response_model=ReviewDetail)
def get_review(
    review_id: uuid.UUID,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Vue détaillée : tous les items groupés par catégorie, unset compris."""
    review = _get_review_or_404(db, review_id)
    _assert_can_view(review, user, db)
    return review_detail(db, review)


# --------------------------------------------------------------------------
# Renommer / changer de statut
# --------------------------------------------------------------------------
@router.patch("/{review_id}", response_model=ReviewDetail)
def update_review(
    review_id: uuid.UUID,
    payload: ReviewUpdate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    review = _get_review_or_404(db, review_id)
    _assert_is_author(review, user)
    if payload.report_name is not None:
        review.report_name = payload.report_name
    if payload.status is not None:
        svc.set_status(db, review=review, new_status=payload.status)
    db.commit()
    db.refresh(review)
    return review_detail(db, review)


# --------------------------------------------------------------------------
# Supprimer
# --------------------------------------------------------------------------
@router.delete("/{review_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_review(
    review_id: uuid.UUID,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    review = _get_review_or_404(db, review_id)
    # Auteur ou admin.
    if review.author_id != user.id and user.role != UserRole.admin:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "Suppression réservée à l'auteur ou à un admin."
        )
    db.delete(review)
    db.commit()


# --------------------------------------------------------------------------
# Mettre à jour un item
# --------------------------------------------------------------------------
@router.patch("/{review_id}/items/{item_id}", response_model=ReviewItemOut)
def update_item(
    review_id: uuid.UUID,
    item_id: uuid.UUID,
    payload: ReviewItemUpdate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Met à jour un point de contrôle puis recalcule le score de la revue."""
    review = _get_review_or_404(db, review_id)
    _assert_is_author(review, user)

    item = db.get(ReviewItem, item_id)
    if item is None or item.review_id != review.id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Point de contrôle introuvable.")

    svc.update_item(db, review=review, item=item, patch=payload.model_dump(exclude_unset=True))
    db.commit()

    # Recharge la vue enrichie de l'item (texte de règle, etc.).
    detail = review_detail(db, review)
    for group in detail.groups:
        for it in group.items:
            if it.id == item.id:
                return it
    raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Item introuvable après màj.")


# --------------------------------------------------------------------------
# Export Excel (colonnes cochées)
# --------------------------------------------------------------------------
_STATUS_COLUMNS = [
    ("OK", ItemStatus.ok),
    ("KO", ItemStatus.ko),
    ("Partiel", ItemStatus.partial),
    ("N/A", ItemStatus.na),
]


@router.get("/{review_id}/export")
def export_review(
    review_id: uuid.UUID,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export Excel formaté d'une revue (colonnes OK/KO/Partiel/N-A cochées)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    review = _get_review_or_404(db, review_id)
    _assert_can_view(review, user, db)
    detail = review_detail(db, review)

    wb = Workbook()
    ws = wb.active
    ws.title = "Revue"

    # En-tête synthèse.
    ws["A1"] = "Rapport"
    ws["B1"] = detail.report_name
    ws["A2"] = "Référentiel"
    ws["B2"] = detail.checklist_type.value
    ws["A3"] = "Score de conformité"
    ws["B3"] = f"{detail.compliance_score or 0} %"
    ws["A4"] = "Statut"
    ws["B4"] = detail.status.value
    for r in range(1, 5):
        ws.cell(row=r, column=1).font = Font(bold=True)

    header_row = 6
    headers = ["Catégorie", "Règle", "Sous-points", "Criticité",
               "OK", "KO", "Partiel", "N/A",
               "Progression", "Risque", "Solution proposée",
               "Jours estimés", "Échéance", "Responsable", "Priorité", "DoD", "Commentaire"]
    ws.append([])  # ligne 5 vide
    for _ in range(header_row - ws.max_row - 1):
        ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="454B66")
    header_font = Font(color="FFFFFF", bold=True)
    check_fill = PatternFill("solid", fgColor="C8E6C9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = border

    for group in detail.groups:
        for it in group.items:
            subs_text = "\n".join(f"• {s}" for s in it.subs)
            row_values = [
                group.category_name,
                it.rule_text,
                subs_text,
                it.criticality.value,
                "", "", "", "",           # colonnes de statut, cochées ci-dessous
                it.progress.value,
                it.risk,
                it.proposed_solution,
                it.estimated_days,
                it.target_date.isoformat() if it.target_date else "",
                it.responsible,
                it.priority,
                it.definition_of_done,
                it.comment,
            ]
            ws.append(row_values)
            row = ws.max_row
            # Cocher la colonne de statut correspondante.
            for offset, (_label, st) in enumerate(_STATUS_COLUMNS):
                col = 5 + offset
                cell = ws.cell(row=row, column=col)
                if it.status == st:
                    cell.value = "✓"
                    cell.fill = check_fill
                    cell.alignment = Alignment(horizontal="center")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in (2, 3, 10, 11, 16, 17):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [22, 50, 34, 12, 5, 5, 8, 6, 13, 26, 26, 12, 12, 16, 10, 26, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    safe = "".join(ch for ch in detail.report_name if ch.isalnum() or ch in " -_")[:60].strip() or "revue"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}.xlsx"'},
    )
