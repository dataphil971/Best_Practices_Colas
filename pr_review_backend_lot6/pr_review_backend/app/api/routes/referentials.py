"""
Routes du référentiel (Lot 2).

Lecture du référentiel actif d'un type de checklist, gestion des catégories,
journal d'activité admin et export Excel formaté.

Le référentiel « actif » = règles `status = 'active'` dont la version courante
est `approved`. Les propositions en attente et les règles retirées sont exposées
par des endpoints admin dédiés (voir routes/rules.py).
"""
import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.auth.deps import get_current_user, require_admin
from app.models.user import User
from app.models.category import Category
from app.models.rule import Rule, RuleVersion
from app.models.rule_activity import RuleActivity
from app.models.enums import ChecklistType, LifecycleState, Criticality
from app.schemas.referential import (
    RuleOut,
    CategoryOut,
    CategoryCreate,
    RuleActivityOut,
)
from app.services.serializers import rule_to_out, activity_to_out

router = APIRouter(prefix="/referentials", tags=["référentiel"])

# Ordre de criticité pour le tri (le plus critique en tête).
_CRIT_RANK = {Criticality.blocking: 0, Criticality.recommended: 1, Criticality.optional: 2}


def _parse_type(type_str: str) -> ChecklistType:
    try:
        return ChecklistType(type_str)
    except ValueError:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            f"Référentiel inconnu : '{type_str}'.",
        )


# --------------------------------------------------------------------------
# Catégories
# --------------------------------------------------------------------------
@router.get("/{type}/categories", response_model=list[CategoryOut])
def list_categories(
    type: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ct = _parse_type(type)
    cats = db.scalars(
        select(Category)
        .where(Category.checklist_type == ct)
        .order_by(Category.order_index, Category.name)
    ).all()
    return list(cats)


@router.post(
    "/{type}/categories",
    response_model=CategoryOut,
    status_code=status.HTTP_201_CREATED,
)
def create_category(
    type: str,
    payload: CategoryCreate,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ct = _parse_type(type)
    if payload.checklist_type != ct:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Le type de la catégorie ne correspond pas à l'URL.",
        )
    exists = db.scalar(
        select(Category).where(
            Category.checklist_type == ct, Category.name == payload.name
        )
    )
    if exists:
        raise HTTPException(status.HTTP_409_CONFLICT, "Catégorie déjà existante.")
    cat = Category(
        checklist_type=ct, name=payload.name, order_index=payload.order_index
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return cat


# --------------------------------------------------------------------------
# Référentiel actif (lecture)
# --------------------------------------------------------------------------
def _active_rules(db: Session, ct: ChecklistType) -> list[tuple[Rule, RuleVersion, Category]]:
    """Règles actives + version courante approuvée + catégorie, jointes."""
    rows = db.execute(
        select(Rule, RuleVersion, Category)
        .join(RuleVersion, RuleVersion.rule_id == Rule.id)
        .join(Category, Category.id == Rule.category_id)
        .where(
            Rule.checklist_type == ct,
            Rule.status == "active",
            RuleVersion.is_current.is_(True),
            RuleVersion.lifecycle == LifecycleState.approved,
        )
    ).all()
    return [(r, v, c) for r, v, c in rows]


@router.get("/{type}", response_model=list[RuleOut])
def get_referential(
    type: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    q: str | None = None,
    sort: str = "category",
):
    """
    Référentiel actif d'un type de checklist.

    Paramètres :
      - `q`     : recherche texte sur le libellé et les sous-points ;
      - `sort`  : `category` (défaut) | `recent` | `criticality`.
    """
    ct = _parse_type(type)
    triples = _active_rules(db, ct)

    if q:
        needle = q.lower()

        def matches(v: RuleVersion) -> bool:
            if needle in v.text.lower():
                return True
            return any(needle in str(s).lower() for s in (v.subs or []))

        triples = [(r, v, c) for (r, v, c) in triples if matches(v)]

    if sort == "recent":
        triples.sort(key=lambda t: t[1].created_at, reverse=True)
    elif sort == "criticality":
        triples.sort(key=lambda t: _CRIT_RANK.get(t[1].criticality, 99))
    else:  # 'category'
        triples.sort(key=lambda t: (t[2].order_index, t[2].name, t[1].text.lower()))

    return [rule_to_out(db, r, category_name=c.name) for (r, v, c) in triples]


# --------------------------------------------------------------------------
# Journal d'activité (admin)
# --------------------------------------------------------------------------
@router.get("/{type}/activity", response_model=list[RuleActivityOut])
def get_activity(
    type: str,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
    action: str | None = None,
    actor: uuid.UUID | None = None,
):
    """Journal d'activité du référentiel, filtrable par action et par acteur."""
    ct = _parse_type(type)
    query = select(RuleActivity).where(RuleActivity.checklist_type == ct)
    if action:
        query = query.where(RuleActivity.action == action)
    if actor:
        query = query.where(RuleActivity.actor_id == actor)
    query = query.order_by(RuleActivity.created_at.desc())
    entries = db.scalars(query).all()
    return [activity_to_out(db, e) for e in entries]


# --------------------------------------------------------------------------
# Export Excel formaté
# --------------------------------------------------------------------------
_CRIT_LABEL = {
    Criticality.blocking: "Bloquant",
    Criticality.recommended: "Recommandé",
    Criticality.optional: "Optionnel",
}


@router.get("/{type}/export")
def export_referential(
    type: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export Excel formaté du référentiel actif."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ct = _parse_type(type)
    triples = _active_rules(db, ct)
    triples.sort(key=lambda t: (t[2].order_index, t[2].name, t[1].text.lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Référentiel {ct.value}"[:31]

    headers = ["Catégorie", "Règle", "Sous-points", "Criticité", "Version"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="454B66")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border

    for _r, v, c in triples:
        subs_text = "\n".join(f"• {s}" for s in (v.subs or []))
        ws.append([
            c.name,
            v.text,
            subs_text,
            _CRIT_LABEL.get(v.criticality, v.criticality.value),
            f"v{v.version_number}",
        ])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [26, 60, 40, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"referentiel_{ct.value}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
